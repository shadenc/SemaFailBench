#!/usr/bin/env python3
"""Re-export every sheet of the frozen Excel workbooks to UTF-8 CSV."""

from __future__ import annotations

import csv
import re
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "docs" / "source_csv"

WORKBOOKS = [
    ROOT / "SemaFailBench_Final_Canary_Dataset_v3_FROZEN.xlsx",
    ROOT.parent / "Corrected_SemaFailBench_Literature_Review (1).xlsx",
]


def safe(name: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "_", name).strip("_")[:80]


def main() -> int:
    OUT.mkdir(parents=True, exist_ok=True)
    manifest = []
    for xlsx in WORKBOOKS:
        if not xlsx.exists():
            print(f"SKIP missing {xlsx}")
            continue
        wb = load_workbook(xlsx, data_only=True, read_only=True)
        book_dir = OUT / safe(xlsx.stem)
        book_dir.mkdir(parents=True, exist_ok=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = [["" if v is None else str(v) for v in row] for row in ws.iter_rows(values_only=True)]
            while rows and all(not c.strip() for c in rows[-1]):
                rows.pop()
            if rows:
                max_c = max(
                    (i for i, col in enumerate(zip(*rows)) if any(str(x).strip() for x in col)),
                    default=-1,
                )
                if max_c >= 0:
                    rows = [r[: max_c + 1] for r in rows]
            out_csv = book_dir / f"{safe(sheet_name)}.csv"
            with out_csv.open("w", encoding="utf-8", newline="") as fh:
                csv.writer(fh).writerows(rows)
            headers = rows[0] if rows else []
            manifest.append(
                {
                    "workbook": xlsx.name,
                    "sheet": sheet_name,
                    "csv_path": str(out_csv.relative_to(ROOT)),
                    "n_rows_including_header": len(rows),
                    "n_cols": len(rows[0]) if rows else 0,
                    "headers": " | ".join(headers),
                }
            )
            print(f"{xlsx.name} / {sheet_name} -> {out_csv.relative_to(ROOT)}")
        wb.close()
    man_path = OUT / "_manifest.csv"
    with man_path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(
            fh,
            fieldnames=["workbook", "sheet", "csv_path", "n_rows_including_header", "n_cols", "headers"],
        )
        writer.writeheader()
        writer.writerows(manifest)
    print(f"manifest -> {man_path.relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
